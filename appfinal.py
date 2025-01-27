import streamlit as st
import pandas as pd
import random
import string
import os
from datetime import date, timedelta
import io
import tempfile

# Librerías para PDF, QR, y Plotly
from fpdf import FPDF
import qrcode
import plotly.express as px
import plotly.io as pio
pio.kaleido.scope.default_format = "png"

# --------------------------------------------------------------------------------
# USUARIOS DEMO
# --------------------------------------------------------------------------------
users_data = {
    "admin": {"password": "admin", "role": "admin"},
    "operador": {"password": "123", "role": "operador"},
    "invitado": {"password": "guest", "role": "invitado"},
}

# --------------------------------------------------------------------------------
# RUTAS DE ARCHIVOS
# --------------------------------------------------------------------------------
LOGO_PATH = "logo1.png"               # Ajusta el nombre si tu logo se llama distinto
EXCEL_FILE_LOTO = "candados_data.xlsx"

def main():
    st.set_page_config(page_title="CandApp by Fossil", layout="wide")

    # Manejo de sesión
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
        st.session_state.current_user = None
        st.session_state.role = None

    # DataFrame de candados
    if "candados_df" not in st.session_state:
        st.session_state["candados_df"] = load_loto_excel()

    # DataFrame de precomisionado
    if "itembook_df" not in st.session_state:
        st.session_state["itembook_df"] = generate_itembook()

    # Si no está logueado, mostrar el login
    if not st.session_state.authenticated:
        apply_custom_styles()
        login()
        st.stop()  # Evita que se dibuje el resto
    else:
        # Usuario logueado
        apply_custom_styles()
        st.image(LOGO_PATH, width=250)
        top_menu()

# --------------------------------------------------------------------------------
# LOGIN
# --------------------------------------------------------------------------------
def login():
    st.markdown(
        """
        <div style="text-align: center; margin-top: 50px;">
            <img src="logo.png" alt="Logo" style="width: 150px; margin-bottom: 20px;">
            <h1 style="color: #4dd0e1;">CandApp by FOSSIL</h1>
            <h3 style="color: #80cbc4; margin-bottom: 20px;">Iniciar sesión</h3>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.form("login_form"):
        username = st.text_input("Usuario", placeholder="Ingresa tu usuario")
        password = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
        submitted = st.form_submit_button("Entrar")

        if submitted:
            if username in users_data:
                if password == users_data[username]["password"]:
                    st.session_state.authenticated = True
                    st.session_state.current_user = username
                    st.session_state.role = users_data[username]["role"]
                    st.success("¡Bienvenido!")
                else:
                    st.error("Contraseña incorrecta.")
            else:
                st.error("Usuario no encontrado.")

# --------------------------------------------------------------------------------
# MENÚ PRINCIPAL (Tabs)
# --------------------------------------------------------------------------------
def top_menu():
    tabs = st.tabs(["LOTO", "Precomisionado", "Salir"])

    with tabs[0]:
        show_loto_section()
    with tabs[1]:
        show_precomisionado_section()
    with tabs[2]:
        st.warning("¿Deseas cerrar sesión?")
        if st.button("Cerrar Sesión"):
            st.session_state.authenticated = False
            st.session_state.current_user = None
            st.session_state.role = None
            st.success("Sesión cerrada.")

# --------------------------------------------------------------------------------
# SECCIÓN LOTO
# --------------------------------------------------------------------------------
def show_loto_section():
    st.markdown("<h2 style='text-align:center; color:#4dd0e1;'>Sección LOTO</h2>", unsafe_allow_html=True)
    sub_tabs = st.tabs([
        "Dashboard",
        "Registrar Candado",
        "Editar/Borrar Candado",
        "Generar Reporte Excel/PDF",
        "Usuarios"
    ])

    with sub_tabs[0]:
        show_dashboard()
    with sub_tabs[1]:
        if st.session_state.role in ["admin", "operador"]:
            input_data()
        else:
            st.error("No tienes permiso para Registrar Candado.")

    with sub_tabs[2]:
        if st.session_state.role == "admin":
            edit_or_delete_candado()
        else:
            st.error("Solo un admin puede Editar/Borrar.")

    with sub_tabs[3]:
        if st.session_state.role in ["admin", "operador"]:
            generate_reports()
        else:
            st.error("Solo operador/admin pueden generar reportes.")

    with sub_tabs[4]:
        if st.session_state.role == "admin":
            manage_users()
        else:
            st.error("Solo admin puede administrar usuarios.")

# --------------------------------------------------------------------------------
# DASHBOARD
# --------------------------------------------------------------------------------
def show_dashboard():
    st.markdown("<h1 style='text-align:center; color:#4dd0e1;'>Lockout-Tagout Dashboard</h1>", unsafe_allow_html=True)
    df = st.session_state["candados_df"]

    if not df.empty:
        col1, col2, col3 = st.columns(3)
        with col1:
            st.markdown("### Total Locks")
            st.metric(label="", value=len(df))
        with col2:
            st.markdown("### Activos")
            activos = (df["Estado"] == "Activo").sum()
            st.metric(label="", value=int(activos))
        with col3:
            st.markdown("### Alertas")
            alertas = (df["Valor"] > 200).sum()
            st.metric(label="", value=int(alertas))

        st.markdown("<h2 style='color:#4dd0e1;'>Gráfico de Candados Activos</h2>", unsafe_allow_html=True)
        fig = plot_active_locks(df)
        st.plotly_chart(fig, use_container_width=True)

        st.markdown("<h2 style='color:#4dd0e1;'>Actividad Reciente</h2>", unsafe_allow_html=True)
        df_sorted = df.sort_values("Fecha", ascending=False)
        for _, row in df_sorted.iterrows():
            st.markdown(
                f"<div style='background:#1c2b3a; padding:10px; margin-bottom:10px;'>"
                f"<span style='color:#ffffff;'>"
                f"No. Candado: {row.get('NoCandado','')} | Área: {row.get('Area','')} | "
                f"Estado: {row.get('Estado','')} | Fecha: {row.get('Fecha','')}"
                f"</span></div>",
                unsafe_allow_html=True
            )
    else:
        st.warning("No hay candados registrados.")

def plot_active_locks(df):
    df_copy = df.copy()
    df_copy["Fecha"] = pd.to_datetime(df_copy["Fecha"], errors="coerce")

    df_activos = df_copy[df_copy["Estado"] == "Activo"]
    if not df_activos.empty:
        df_count = df_activos.groupby(df_activos["Fecha"].dt.date).size().reset_index(name="count")
    else:
        df_count = pd.DataFrame({"Fecha": [], "count": []})

    fig = px.line(
        df_count,
        x="Fecha",
        y="count",
        markers=True,
        title="Tendencia de Candados Activos",
    )
    fig.update_layout(
        plot_bgcolor="#1c2b3a",
        paper_bgcolor="#0e1a2b",
        font_color="#ffffff",
        title_font_color="#4dd0e1"
    )
    fig.update_traces(line_color="#4dd0e1", marker_color="#4dd0e1")
    return fig

# --------------------------------------------------------------------------------
# REGISTRAR CANDADO
# --------------------------------------------------------------------------------
def input_data():
    st.markdown("<h1 style='text-align:center; color:#4dd0e1;'>Registrar Nuevo Candado</h1>", unsafe_allow_html=True)
    with st.form("register_lock"):
        no_candado = st.text_input("No. de Candado")
        area = st.text_input("Área")
        tablero_equipo = st.text_input("Tablero o Equipo")
        kks = st.text_input("KKS")
        tipo_bloqueo = st.text_input("Tipo de Bloqueo")
        lider_aut = st.text_input("Líder Autorizador")
        ejecutado_por_nombre = st.text_input("Bloqueo Ejecutado Por - Nombre")
        ejecutado_por_cargo = st.text_input("Bloqueo Ejecutado Por - Cargo")
        n_ptw = st.text_input("N° PTW")
        fecha_reg = st.date_input("Fecha de Bloqueo", value=date.today())

        descripcion = st.text_area("Descripción (opcional)", "")
        responsable = st.text_input("Responsable (opcional)")
        estado_check = st.checkbox("Activo", value=True)
        valor = st.number_input("Valor (opcional)", min_value=0, max_value=99999, value=0)

        uploaded_file = st.file_uploader("Adjuntar PDF (opcional)", type=["pdf"])

        submitted = st.form_submit_button("Guardar Registro")

        if submitted:
            pdf_data = None
            if uploaded_file is not None:
                file_bytes = uploaded_file.read()
                pdf_data = bytes(file_bytes)  # Forzar a bytes

            # Generar QR
            data_qr = f"NoCandado={no_candado}, Area={area}, Fecha={fecha_reg}"
            qr_bytes = generate_qr_code(data_qr)

            df = st.session_state["candados_df"]
            new_row = {
                "ID": no_candado,
                "NoCandado": no_candado,
                "Area": area,
                "TableroEquipo": tablero_equipo,
                "KKS": kks,
                "TipoBloqueo": tipo_bloqueo,
                "LiderAutorizador": lider_aut,
                "EjecPorNombre": ejecutado_por_nombre,
                "EjecPorCargo": ejecutado_por_cargo,
                "N_PTW": n_ptw,
                "Fecha": str(fecha_reg),
                "Descripción": descripcion,
                "Responsable": responsable,
                "Estado": "Activo" if estado_check else "Inactivo",
                "Valor": valor,
                "QR_Bytes": qr_bytes,
                "PDF_Adjunto": pdf_data
            }

            new_df = pd.DataFrame([new_row])
            df = pd.concat([df, new_df], ignore_index=True)

            st.session_state["candados_df"] = df
            save_loto_excel(df)
            st.success("Registro guardado exitosamente.")

def generate_qr_code(data: str) -> bytes:
    qr = qrcode.QRCode(version=1, box_size=5, border=4)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf.getvalue()

# --------------------------------------------------------------------------------
# GENERAR TARJETA DE PELIGRO (PDF)
# --------------------------------------------------------------------------------
def generate_loto_card(row) -> bytes:
    """
    Genera un PDF estilo "Tarjeta de Peligro LOTO" con fondo rojo y texto blanco.
    Incluye QR (si existe), y fuerza salida en bytes para evitar bytearray.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=5)

    # Fondo rojo
    pdf.set_fill_color(255, 0, 0)
    page_width, page_height = pdf.w, pdf.h
    pdf.rect(5, 5, page_width - 10, page_height - 10, style='F')

    # Texto blanco
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Arial", "B", 18)
    pdf.set_xy(10, 15)
    pdf.cell(0, 10, "PELIGRO - LOTO", ln=1, align="C")
    pdf.ln(5)

    # Datos
    pdf.set_font("Arial", "", 12)
    def centered_line(text):
        pdf.cell(0, 8, text, ln=1, align="C")

    no_candado = row.get('NoCandado', '')
    area = row.get('Area', '')
    tablero = row.get('TableroEquipo', '')
    tipo = row.get('TipoBloqueo', '')
    fecha = row.get('Fecha', '')

    centered_line(f"No. de Candado: {no_candado}")
    centered_line(f"Área: {area}")
    centered_line(f"Equipo/Tablero: {tablero}")
    centered_line(f"Tipo: {tipo}")
    centered_line(f"Fecha: {fecha}")

    # Insertar QR si existe
    qr_data = row.get("QR_Bytes", None)
    if qr_data and isinstance(qr_data, (bytes, bytearray)):
        qr_data = bytes(qr_data)  # Convertir a bytes si bytearray
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            tmp.write(qr_data)
            tmp.flush()
            current_y = pdf.get_y() + 10
            x_qr = (page_width - 30) / 2
            pdf.image(tmp.name, x=x_qr, y=current_y, w=30)

    # Forzar la salida a bytes
    pdf_bytes = pdf.output(dest="S")
    return bytes(pdf_bytes)

# --------------------------------------------------------------------------------
# EDITAR/BORRAR CANDADO (con botón Generar Tarjeta)
# --------------------------------------------------------------------------------
def edit_or_delete_candado():
    st.markdown("<h1 style='text-align:center; color:#4dd0e1;'>Editar o Borrar Candados</h1>", unsafe_allow_html=True)
    df = st.session_state["candados_df"]
    if df.empty:
        st.info("No hay candados para editar/borrar.")
        return

    st.subheader("Candados disponibles:")
    df_display = df[["NoCandado", "Area", "Estado", "Fecha"]].reset_index(drop=True)
    if df_display.empty:
        st.warning("No hay datos.")
        return

    if "edit_mode" not in st.session_state:
        st.session_state["edit_mode"] = None
    if "tarjeta_pdf" not in st.session_state:
        st.session_state["tarjeta_pdf"] = None
    if "tarjeta_idx" not in st.session_state:
        st.session_state["tarjeta_idx"] = None

    select_idx = st.selectbox(
        "Elige un candado para editar/borrar/generar tarjeta:",
        options=df_display.index,
        format_func=lambda i: f"No. {df_display.loc[i, 'NoCandado']} | Área: {df_display.loc[i, 'Area']}"
    )
    row_data = df.iloc[select_idx]

    col1, col2, col3 = st.columns([1,1,1])
    with col1:
        if st.button("Editar"):
            st.session_state["edit_mode"] = select_idx
    with col2:
        if st.button("Borrar"):
            df.drop(df.index[select_idx], inplace=True)
            st.session_state["candados_df"] = df
            save_loto_excel(df)
            st.success("Candado borrado.")
    with col3:
        if st.button("Generar Tarjeta"):
            # Genera la tarjeta y la almacena en session_state
            pdf_card = generate_loto_card(row_data)
            st.session_state["tarjeta_pdf"] = pdf_card
            st.session_state["tarjeta_idx"] = select_idx
            st.success("Tarjeta generada.")

    # Mostrar botón de descarga si tenemos la tarjeta generada para este candado
    if (
        st.session_state["tarjeta_pdf"] is not None
        and st.session_state["tarjeta_idx"] == select_idx
    ):
        # "data" debe ser bytes, no bytearray
        data_bytes = bytes(st.session_state["tarjeta_pdf"])
        st.download_button(
            label="Descargar Tarjeta PDF",
            data=data_bytes,
            file_name=f"tarjeta_{row_data.get('NoCandado','')}.pdf",
            mime="application/pdf"
        )

    # Si estamos en modo edición
    if st.session_state["edit_mode"] == select_idx:
        with st.expander(f"Editando No. Candado: {row_data.get('NoCandado','')}", expanded=True):
            edit_candado_form(select_idx)

def edit_candado_form(idx):
    df = st.session_state["candados_df"]
    candado = df.loc[idx]

    with st.form(f"edit_form_{idx}", clear_on_submit=True):
        no_candado = st.text_input("No. de Candado", value=candado.get("NoCandado", ""))
        area = st.text_input("Área", value=candado.get("Area",""))
        tablero_equipo = st.text_input("Tablero o Equipo", value=candado.get("TableroEquipo",""))
        kks = st.text_input("KKS", value=candado.get("KKS",""))
        tipo_bloqueo = st.text_input("Tipo de Bloqueo", value=candado.get("TipoBloqueo",""))
        lider_aut = st.text_input("Líder Autorizador", value=candado.get("LiderAutorizador",""))
        e_nom = st.text_input("Bloqueo Ejecutado Por - Nombre", value=candado.get("EjecPorNombre",""))
        e_cargo = st.text_input("Bloqueo Ejecutado Por - Cargo", value=candado.get("EjecPorCargo",""))
        n_ptw = st.text_input("N° PTW", value=candado.get("N_PTW",""))

        fecha_str = candado.get("Fecha", str(date.today()))
        try:
            fecha_val = pd.to_datetime(fecha_str).date()
        except:
            fecha_val = date.today()
        new_fecha = st.date_input("Fecha", value=fecha_val)

        new_desc = st.text_area("Descripción", value=candado.get("Descripción",""))
        new_resp = st.text_input("Responsable", value=candado.get("Responsable",""))
        estado_actual = candado.get("Estado", "Activo")
        new_estado = st.selectbox("Estado", ["Activo", "Inactivo"],
                                  index=0 if estado_actual == "Activo" else 1)
        new_valor = st.number_input("Valor", min_value=0, max_value=999999,
                                    value=int(candado.get("Valor", 0)))

        submitted = st.form_submit_button("Guardar Cambios")
        if submitted:
            df.at[idx, "NoCandado"] = no_candado
            df.at[idx, "Area"] = area
            df.at[idx, "TableroEquipo"] = tablero_equipo
            df.at[idx, "KKS"] = kks
            df.at[idx, "TipoBloqueo"] = tipo_bloqueo
            df.at[idx, "LiderAutorizador"] = lider_aut
            df.at[idx, "EjecPorNombre"] = e_nom
            df.at[idx, "EjecPorCargo"] = e_cargo
            df.at[idx, "N_PTW"] = n_ptw
            df.at[idx, "Fecha"] = str(new_fecha)
            df.at[idx, "Descripción"] = new_desc
            df.at[idx, "Responsable"] = new_resp
            df.at[idx, "Estado"] = new_estado
            df.at[idx, "Valor"] = new_valor
            df.at[idx, "ID"] = no_candado

            # Regenerar QR
            data_qr = f"NoCandado={no_candado}, Area={area}, Fecha={new_fecha}"
            df.at[idx, "QR_Bytes"] = generate_qr_code(data_qr)

            st.session_state["candados_df"] = df
            save_loto_excel(df)
            st.success("Cambios guardados.")
            # Salir del modo edición
            st.session_state["edit_mode"] = None

# --------------------------------------------------------------------------------
# GENERAR REPORTES (Excel / PDF)
# --------------------------------------------------------------------------------
def generate_reports():
    st.markdown("<h1 style='text-align:center; color:#4dd0e1;'>Generar Reporte Excel / PDF</h1>", unsafe_allow_html=True)
    df = st.session_state["candados_df"]
    if df.empty:
        st.warning("No hay datos para exportar.")
        return

    formato = st.radio("Selecciona formato de reporte:", ["Excel", "PDF"], horizontal=True)
    if formato == "Excel":
        if st.button("Generar y Descargar Excel"):
            excel_bytes = generate_excel_file(df)
            st.download_button(
                label="Descargar Reporte en Excel",
                data=excel_bytes,
                file_name="reporte_candados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.write("Puedes generar un PDF con uno o con todos los candados.")
        candados_list = ["Todos"] + df["NoCandado"].dropna().unique().tolist()
        selected = st.selectbox("Selecciona un candado específico o 'Todos':", candados_list)

        if st.button("Generar y Descargar PDF"):
            if selected == "Todos":
                pdf_bytes = generate_pdf_all(df)
            else:
                sub_df = df[df["NoCandado"] == selected]
                pdf_bytes = generate_pdf_all(sub_df)

            # Forzar a bytes en caso de bytearray
            pdf_bytes = bytes(pdf_bytes)
            st.download_button(
                label="Descargar Reporte en PDF",
                data=pdf_bytes,
                file_name="candados.pdf",
                mime="application/pdf"
            )

def generate_excel_file(df: pd.DataFrame) -> bytes:
    import openpyxl
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Excluir QR_Bytes y PDF_Adjunto
        cols = [c for c in df.columns if c not in ["QR_Bytes", "PDF_Adjunto"]]
        df[cols].to_excel(writer, index=False, sheet_name="ReporteCandados")
    output.seek(0)
    return output.getvalue()

def generate_pdf_all(df: pd.DataFrame) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Arial", size=11)

    for _, row in df.iterrows():
        pdf.add_page()

        # Insertar LOGO centrado
        page_width = pdf.w
        logo_w = 50
        x_logo = (page_width - logo_w) / 2
        pdf.image(LOGO_PATH, x=x_logo, y=10, w=logo_w)
        pdf.ln(35)

        pdf.cell(0, 5, txt=f"No. de Candado: {row.get('NoCandado','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Área: {row.get('Area','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Tablero/Equipo: {row.get('TableroEquipo','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"KKS: {row.get('KKS','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Tipo de Bloqueo: {row.get('TipoBloqueo','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Líder Autorizador: {row.get('LiderAutorizador','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Ejecutado por: {row.get('EjecPorNombre','')} (Cargo: {row.get('EjecPorCargo','')})", ln=1, align="C")
        pdf.cell(0, 5, txt=f"N° PTW: {row.get('N_PTW','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Fecha: {row.get('Fecha','')}", ln=1, align="C")

        pdf.ln(5)
        pdf.cell(0, 5, txt=f"Descripción: {row.get('Descripción','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Responsable: {row.get('Responsable','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Estado: {row.get('Estado','')}", ln=1, align="C")
        pdf.cell(0, 5, txt=f"Valor: {row.get('Valor',0)}", ln=1, align="C")

        # Inserta el QR
        qr_data = row.get("QR_Bytes", None)
        if qr_data and isinstance(qr_data, (bytes, bytearray)):
            qr_data = bytes(qr_data)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp.write(qr_data)
                tmp.flush()
                pdf.image(tmp.name, x=(page_width - 30)/2, y=pdf.get_y() + 5, w=30)

        pdf.ln(40)
        pdf.cell(0, 5, txt="---------------------------------", ln=1, align="C")

    pdf_bytes = pdf.output(dest="S")
    return bytes(pdf_bytes)

# --------------------------------------------------------------------------------
# ADMINISTRAR USUARIOS
# --------------------------------------------------------------------------------
def manage_users():
    st.markdown("<h1 style='text-align:center; color:#4dd0e1;'>Administrar Usuarios</h1>", unsafe_allow_html=True)

    st.subheader("Usuarios actuales:")
    for user, info in users_data.items():
        st.write(f"- **{user}** (rol: {info['role']})")

    st.write("---")
    st.subheader("Crear nuevo usuario")
    with st.form("new_user_form"):
        new_username = st.text_input("Nombre de usuario")
        new_password = st.text_input("Contraseña", type="password")
        new_role = st.selectbox("Rol", ["admin", "operador", "invitado"])
        create_submitted = st.form_submit_button("Crear Usuario")
        if create_submitted:
            if new_username in users_data:
                st.error("Ese usuario ya existe.")
            else:
                users_data[new_username] = {"password": new_password, "role": new_role}
                st.success(f"Usuario '{new_username}' creado con rol '{new_role}'.")

# --------------------------------------------------------------------------------
# PRECOMISIONADO (NO MODIFICAR)
# --------------------------------------------------------------------------------
def show_precomisionado_section():
    st.markdown("<h2 style='text-align:center; color:#4dd0e1;'>Precomisionado - Dossier Digital</h2>", unsafe_allow_html=True)

    sub_tabs = st.tabs(["Items", "Generar ITR (PDF)", "Formulario Excel Dinámico"])
    with sub_tabs[0]:
        show_item_list()
    with sub_tabs[1]:
        generate_itr_pdf()
    with sub_tabs[2]:
        run_document_form()

def show_item_list():
    st.write("**Items** (ejemplo) para Precomisionado:")
    df_items = st.session_state["itembook_df"]
    st.dataframe(df_items)

def generate_itr_pdf():
    st.write("Completa el formulario de ITR y genera un PDF similar al ejemplo.")
    df_items = st.session_state["itembook_df"]
    if df_items.empty:
        st.warning("No hay items en la base de datos.")
        return

    item_id = st.selectbox("Seleccionar ItemID", df_items["ItemID"].unique())
    row_item = df_items[df_items["ItemID"] == item_id].iloc[0]

    with st.form("itr_form"):
        equipo = st.text_input("Descripción del Equipo", value=row_item["Descripcion"])
        subsistema = st.text_input("Sub-sistema", "Tensión segura (ejemplo)")
        responsable = st.text_input("Responsable", "Ing. Precomisionado")
        comentarios = st.text_area("Comentarios", "Observaciones...")
        submitted = st.form_submit_button("Generar PDF")

    if submitted:
        pdf_bytes = generar_pdf_precom(row_item, equipo, subsistema, responsable, comentarios)
        st.session_state["pdf_bytes"] = pdf_bytes
        st.success("PDF generado con éxito. Descarga a continuación:")

    if "pdf_bytes" in st.session_state and st.session_state["pdf_bytes"] is not None:
        st.download_button(
            label="Descargar ITR PDF",
            data=st.session_state["pdf_bytes"],
            file_name=f"ITR_{item_id}.pdf",
            mime="application/pdf"
        )

def generar_pdf_precom(item_row, equipo, subsistema, responsable, comentarios):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "", 12)

    pdf.cell(0, 10, txt="E11A - Centro de Control de Motores (BT/AT) (MCC)", ln=1, align="C")
    pdf.cell(0, 10, txt="Completamiento de la Construcción", ln=1, align="C")

    pdf.ln(5)
    pdf.cell(0, 8, txt=f"N° de Tag: {item_row['ItemID']}", ln=1)
    pdf.cell(0, 8, txt=f"Descripción del Equipo: {equipo}", ln=1)
    pdf.cell(0, 8, txt=f"N° de Subsistema: {subsistema}", ln=1)
    pdf.cell(0, 8, txt=f"Proyecto: {item_row['Proyecto']}", ln=1)
    pdf.cell(0, 8, txt=f"Responsable: {responsable}", ln=1)

    pdf.ln(5)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, txt="Items para verificar:", ln=1)
    pdf.set_font("Arial", "", 12)
    pdf.multi_cell(0, 8, "- Placa de identificación\n- Dispositivo de fijación\n- MCCB, contactores...")

    pdf.ln(5)
    pdf.cell(0, 8, txt="Comentarios / Observaciones:", ln=1)
    pdf.multi_cell(0, 8, comentarios)
    pdf.ln(10)
    pdf.cell(0, 8, txt="Firmado por: _______________________", ln=1)
    pdf.cell(0, 8, txt="Fecha: _____________________________", ln=1)

    pdf_bytes = pdf.output(dest="S")
    return bytes(pdf_bytes)  # Forzar a bytes

def run_document_form():
    st.write("### Crear formulario a partir de un archivo Excel")
    uploaded_file = st.file_uploader("Subir archivo Excel", type=["xlsx", "xls"])
    if uploaded_file is None:
        st.info("Por favor, sube un archivo para continuar.")
        return

    try:
        df_def = pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Error al leer el Excel: {e}")
        return

    st.write("Se generará un formulario basado en las filas del Excel.")
    st.write("Columnas esperadas: field_type, label, options, default (puede variar)")

    form_values = {}

    with st.form("dynamic_form"):
        for i, row in df_def.iterrows():
            field_type = str(row.get("field_type", "")).lower()
            label = row.get("label", f"Campo {i}")
            options = row.get("options", "")
            default = row.get("default", "")

            if field_type == "text":
                form_values[label] = st.text_input(label, value=str(default))
            elif field_type == "checkbox":
                default_bool = (str(default).lower() == "true")
                form_values[label] = st.checkbox(label, value=default_bool)
            elif field_type == "select":
                opt_list = [o.strip() for o in str(options).split(",")]
                if default not in opt_list:
                    default_index = 0
                else:
                    default_index = opt_list.index(default)
                form_values[label] = st.selectbox(label, opt_list, index=default_index)
            else:
                form_values[label] = st.text_input(label, value=str(default))

        submitted = st.form_submit_button("Generar PDF")

    if submitted:
        pdf_bytes = generar_pdf_dinamico(form_values)
        st.success("Se generó el PDF con la información. Descarga abajo:")

        st.download_button(
            label="Descargar PDF",
            data=pdf_bytes,
            file_name="formulario_generado.pdf",
            mime="application/pdf"
        )

def generar_pdf_dinamico(form_data: dict) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "", 12)

    pdf.cell(0, 10, txt="Formulario Dinámico - Resultado", ln=True, align="C")
    pdf.ln(5)

    for label, value in form_data.items():
        pdf.multi_cell(0, 8, f"{label}: {value}")
        pdf.ln(2)

    pdf_bytes = pdf.output(dest="S")
    return bytes(pdf_bytes)  # Forzar a bytes

# --------------------------------------------------------------------------------
# LECTURA/ESCRITURA DE CANDADOS
# --------------------------------------------------------------------------------
def load_loto_excel():
    if not os.path.exists(EXCEL_FILE_LOTO):
        df = prepopulate_loto(n=30)
        df.to_excel(EXCEL_FILE_LOTO, index=False)
    else:
        df = pd.read_excel(EXCEL_FILE_LOTO)
        needed_cols = [
            "NoCandado","Area","TableroEquipo","KKS","TipoBloqueo","LiderAutorizador",
            "EjecPorNombre","EjecPorCargo","N_PTW","QR_Bytes","PDF_Adjunto","Valor",
            "Estado","Descripción","Responsable","Fecha","ID"
        ]
        for col in needed_cols:
            if col not in df.columns:
                df[col] = ""
    return df

def save_loto_excel(df):
    df.to_excel(EXCEL_FILE_LOTO, index=False)

def prepopulate_loto(n=30):
    rows = []
    today = date.today()
    for i in range(n):
        candado_id = f"Rojo{i+1}"
        area = random.choice(["SHELTER LV", "Sala Compresores", "Tanques", "Area Baterías"])
        tablero = random.choice(["UPS", "UPS DISTRIBUTION BOARD", "Q74", "Q43"])
        kks_val = random.choice(["Q73", "Q74", "Q43", "Q99"])
        tipo = f"CANDADO {i+1}"
        lider = random.choice(["Monsu Ariel", "Avecilla Miguel", "Scimeca Gabriel"])
        ejecutor = random.choice(["Perez Martin", "Sanchez Pedro", "Lopez Carlos"])
        cargo = random.choice(["Supervisor", "Operador", "Técnico"])
        ptw_number = str(random.randint(1,10))
        days_back = random.randint(0, 60)
        fecha_rand = today - timedelta(days=days_back)
        estado = random.choice(["Activo", "Inactivo"])

        qr_str = f"NoCandado={candado_id}, Area={area}, Fecha={fecha_rand}"
        qr_bytes = generate_qr_code(qr_str)

        row = {
            "ID": candado_id,
            "NoCandado": candado_id,
            "Area": area,
            "TableroEquipo": tablero,
            "KKS": kks_val,
            "TipoBloqueo": tipo,
            "LiderAutorizador": lider,
            "EjecPorNombre": ejecutor,
            "EjecPorCargo": cargo,
            "N_PTW": ptw_number,
            "Fecha": str(fecha_rand),
            "Descripción": f"Descripción {i+1}",
            "Responsable": lider,
            "Estado": estado,
            "Valor": random.randint(0, 300),
            "QR_Bytes": qr_bytes,
            "PDF_Adjunto": None,
        }
        rows.append(row)
    return pd.DataFrame(rows)

# --------------------------------------------------------------------------------
# ÍTEMS DE PRECOMISIONADO
# --------------------------------------------------------------------------------
def generate_itembook():
    proyectos = ["Proyecto A"] * 20 + ["Proyecto B"] * 20
    rows = []
    for i in range(40):
        item_id = f"ITM-{i+1:03d}"
        desc = "Item Ejemplo " + "".join(random.choices(string.ascii_uppercase, k=2))
        row = {
            "Proyecto": proyectos[i],
            "ItemID": item_id,
            "Descripcion": desc,
        }
        rows.append(row)
    return pd.DataFrame(rows)

# --------------------------------------------------------------------------------
# APLICAR ESTILOS (CSS)
# --------------------------------------------------------------------------------
def apply_custom_styles():
    st.markdown(
        """
        <style>
            [data-testid="stAppViewContainer"] {
                background-color: #0e1a2b !important;
            }
            [data-testid="stHeader"] {
                background-color: #0e1a2b !important;
            }
            html, body, [class*="css"]  {
                color: #ffffff !important;
            }
            .stTabs [role="tablist"] button [data-baseweb="tab"] {
                color: #ffffff !important;
                background-color: #0e1a2b !important;
                border: 1px solid #4dd0e1 !important;
            }
            .stTabs [role="tablist"] button[aria-selected="true"] {
                background-color: #1c2b3a !important;
                color: #4dd0e1 !important;
            }
        </style>
        """,
        unsafe_allow_html=True
    )

# --------------------------------------------------------------------------------
# EJECUTAR
# --------------------------------------------------------------------------------
if __name__ == "__main__":
    main()
